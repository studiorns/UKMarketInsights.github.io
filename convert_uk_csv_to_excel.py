import pandas as pd
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def convert_csv_to_excel(csv_path, excel_path):
    """
    Convert the UK Travel Queries Forecast CSV file to Excel format.
    
    Args:
        csv_path (str): Path to the CSV file
        excel_path (str): Path where the Excel file will be saved
    """
    print(f"Converting {csv_path} to Excel format...")
    
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    
    # Read the CSV file
    with open(csv_path, 'r') as file:
        content = file.read()
    
    # Split the content by section markers
    sections = content.split('====================')
    
    # Process each section
    for i, section in enumerate(sections):
        if not section.strip():
            continue
        
        # Extract section name and content
        lines = section.strip().split('\n')
        
        # For the first section (header), use "Introduction" as the sheet name
        if i == 0:
            sheet_name = "Introduction"
            section_content = lines
        else:
            # The first line after the section marker is the section name
            section_name = lines[0].strip()
            section_content = lines[1:]
            sheet_name = section_name.replace(" ", "_")[:31]  # Excel has a 31 character limit for sheet names
        
        # Create a new sheet
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Add the content to the sheet
        for row_idx, line in enumerate(section_content):
            if line.strip():  # Skip empty lines
                cells = line.split(',')
                for col_idx, cell_value in enumerate(cells):
                    sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value.strip())
        
        # Format the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
        header_alignment = Alignment(wrap_text=True, vertical="top")
        header_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Set column widths
        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Add a summary sheet
    summary_sheet = workbook.create_sheet(title="Summary")
    
    # Add summary data
    summary_data = [
        ["Metric", "Value"],
        ["Average 2024 Monthly Queries", "97.56"],
        ["Conservative Forecast Average", "103.78"],
        ["Moderate Forecast Average", "109.47"],
        ["Ambitious Forecast Average", "115.16"],
        ["Conservative YoY Growth", "6.4%"],
        ["Moderate YoY Growth", "12.2%"],
        ["Ambitious YoY Growth", "18.0%"],
        ["Peak Month", "October"],
        ["Highest Growth Month", "October"]
    ]
    
    for row_idx, row_data in enumerate(summary_data):
        for col_idx, cell_value in enumerate(row_data):
            summary_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
    
    # Format the summary header
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Set column widths for summary
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    
    # Save the workbook
    workbook.save(excel_path)
    
    print(f"Conversion complete. Excel file saved to {excel_path}")

def main():
    # Get the directory of the current script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # List of CSV files to convert
    csv_files = [
        "Travel_Queries_Forecast_UK_Updated.csv",
        "Travel_Queries_Forecast_UK_Conservative.csv",
        "Travel_Queries_Forecast_UK_Enhanced.csv",
        "Travel_Queries_Forecast_UK_Enhanced_Updated.csv",
        "UK_Travel_Queries_2025_Forecast_Updated.csv"
    ]
    
    # Convert each CSV file to Excel
    for csv_file in csv_files:
        csv_path = os.path.join(script_dir, csv_file)
        excel_path = os.path.join(script_dir, csv_file.replace(".csv", ".xlsx"))
        
        if os.path.exists(csv_path):
            convert_csv_to_excel(csv_path, excel_path)
            print(f"{csv_file} conversion complete.")
        else:
            print(f"File not found: {csv_path}")

if __name__ == "__main__":
    main()
